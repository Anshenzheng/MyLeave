from flask import Blueprint, request, jsonify, send_file
from flask_login import login_required, current_user
from app import db
from app.models import LeaveApplication, User, Department, HolidayType
from datetime import datetime, timedelta, date
from dateutil import parser
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import io
import tempfile

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/calendar', methods=['GET'])
@login_required
def get_calendar_data():
    year = request.args.get('year', type=int, default=datetime.now().year)
    month = request.args.get('month', type=int)
    department_id = request.args.get('department_id', type=int)

    if current_user.role == 'employee':
        return jsonify({'error': '权限不足'}), 403

    start_date = date(year, month, 1) if month else date(year, 1, 1)
    if month:
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, 12, 31)

    query = LeaveApplication.query.filter(
        LeaveApplication.status == 'approved',
        LeaveApplication.start_date <= end_date,
        LeaveApplication.end_date >= start_date
    )

    if current_user.role == 'manager':
        if department_id:
            query = query.join(User).filter(User.department_id == department_id)
        else:
            query = query.join(User).filter(User.department_id == current_user.department_id)
    elif department_id:
        query = query.join(User).filter(User.department_id == department_id)

    applications = query.all()

    calendar_data = {}
    for app in applications:
        current = max(app.start_date, start_date)
        while current <= min(app.end_date, end_date):
            date_str = current.isoformat()
            if date_str not in calendar_data:
                calendar_data[date_str] = []
            calendar_data[date_str].append({
                'id': app.id,
                'user_id': app.user_id,
                'employee_name': app.employee.name if app.employee else None,
                'employee_department': app.employee.department.name if app.employee and app.employee.department else None,
                'holiday_type_name': app.holiday_type.name if app.holiday_type else None,
                'start_date': app.start_date.isoformat(),
                'end_date': app.end_date.isoformat()
            })
            current += timedelta(days=1)

    return jsonify({'calendar_data': calendar_data}), 200

@reports_bp.route('/export', methods=['GET'])
@login_required
def export_excel():
    if current_user.role not in ['admin', 'manager']:
        return jsonify({'error': '权限不足'}), 403

    department_id = request.args.get('department_id', type=int)
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    status = request.args.get('status', 'approved')

    query = LeaveApplication.query

    if current_user.role == 'manager':
        if department_id:
            query = query.join(User).filter(User.department_id == department_id)
        else:
            query = query.join(User).filter(User.department_id == current_user.department_id)
    elif department_id:
        query = query.join(User).filter(User.department_id == department_id)

    if start_date_str:
        query = query.filter(LeaveApplication.start_date >= parser.parse(start_date_str).date())
    if end_date_str:
        query = query.filter(LeaveApplication.end_date <= parser.parse(end_date_str).date())
    if status:
        query = query.filter_by(status=status)

    applications = query.order_by(LeaveApplication.start_date.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "请假记录"

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ['序号', '员工姓名', '部门', '假期类型', '开始日期', '结束日期', '天数', '请假原因', '状态', '审批人', '审批时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    status_map = {
        'pending': '待审批',
        'approved': '已通过',
        'rejected': '已拒绝',
        'cancelled': '已取消'
    }

    for row, app in enumerate(applications, 2):
        ws.cell(row=row, column=1, value=row - 1).border = thin_border
        ws.cell(row=row, column=2, value=app.employee.name if app.employee else '').border = thin_border
        ws.cell(row=row, column=3, value=app.employee.department.name if app.employee and app.employee.department else '').border = thin_border
        ws.cell(row=row, column=4, value=app.holiday_type.name if app.holiday_type else '').border = thin_border
        ws.cell(row=row, column=5, value=app.start_date.isoformat() if app.start_date else '').border = thin_border
        ws.cell(row=row, column=6, value=app.end_date.isoformat() if app.end_date else '').border = thin_border
        ws.cell(row=row, column=7, value=app.days).border = thin_border
        ws.cell(row=row, column=8, value=app.reason or '').border = thin_border
        ws.cell(row=row, column=9, value=status_map.get(app.status, app.status)).border = thin_border
        ws.cell(row=row, column=10, value=app.approver.name if app.approver else '').border = thin_border
        ws.cell(row=row, column=11, value=app.approved_at.strftime('%Y-%m-%d %H:%M') if app.approved_at else '').border = thin_border

    for col in range(1, 12):
        ws.column_dimensions[chr(64 + col)].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'请假记录_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@reports_bp.route('/statistics', methods=['GET'])
@login_required
def get_statistics():
    if current_user.role not in ['admin', 'manager']:
        return jsonify({'error': '权限不足'}), 403

    year = request.args.get('year', type=int, default=datetime.now().year)
    department_id = request.args.get('department_id', type=int)

    query = LeaveApplication.query.filter(
        LeaveApplication.status == 'approved'
    ).join(User)

    if current_user.role == 'manager':
        if department_id:
            query = query.filter(User.department_id == department_id)
        else:
            query = query.filter(User.department_id == current_user.department_id)
    elif department_id:
        query = query.filter(User.department_id == department_id)

    applications = query.all()

    by_type = {}
    by_department = {}
    total_days = 0

    for app in applications:
        app_year = app.start_date.year
        if app_year != year:
            continue

        type_name = app.holiday_type.name if app.holiday_type else '未知'
        if type_name not in by_type:
            by_type[type_name] = {'count': 0, 'days': 0}
        by_type[type_name]['count'] += 1
        by_type[type_name]['days'] += app.days

        dept_name = app.employee.department.name if app.employee and app.employee.department else '未分配'
        if dept_name not in by_department:
            by_department[dept_name] = {'count': 0, 'days': 0}
        by_department[dept_name]['count'] += 1
        by_department[dept_name]['days'] += app.days

        total_days += app.days

    return jsonify({
        'year': year,
        'total_applications': len([a for a in applications if a.start_date.year == year]),
        'total_days': total_days,
        'by_type': by_type,
        'by_department': by_department
    }), 200
